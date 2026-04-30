#!/usr/bin/env python3
"""Generate RED EN 18031 Excel workbooks from ASUS router templates.

The EN 18031-1 template is already an .xlsx file, so openpyxl can preserve its
formatting and formulas. The EN 18031-2 source supplied for this workflow is an
.xlsb file; pyxlsb can read its values but cannot write the original binary
workbook format, so this script exports it as a value-based .xlsx workbook.
"""

from __future__ import annotations

import argparse
import re
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pyxlsb import open_workbook as open_xlsb


DEFAULT_EN18031_1_TEMPLATE = Path(
    "/Users/wilsondenq879/Downloads/"
    "RED-DA_ EN-18031-1_Conceptual-Assessment_ Decision-ASUS Router_[Model name].xlsx"
)
DEFAULT_EN18031_2_TEMPLATE = Path(
    "/Users/wilsondenq879/Downloads/RED 18031-2 ASUS Router_[Model name].xlsb"
)

WIFI_CHOICES = {
    "2.4-5": "Wi-Fi Radio (2.4 & 5GHz)",
    "2.4-5-6": "Wi-Fi Radio (2.4 & 5 & 6GHz)",
}

OLD_UM = "E27527_RT-BE59_UM_EAA"
OLD_QSG = "U26937_RT-BE59_folded_EU_QSG_312x450mm_PRINT"
OLD_SPEC = "https://www.asus.com/networking-iot-servers/wifi-7/all-series/rt-be59/techspec/"

WIFI_PATTERN = re.compile(r"Wi-Fi Radio \((?:2\.4\s*(?:&|,)\s*5|2\.4\s*(?:&|,)\s*5\s*(?:&|,)\s*6)\s*GHz\)", re.I)
ILLEGAL_XLSX_CHAR_PATTERN = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def timestamp() -> str:
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def normalize_wifi_choice(value: str) -> str:
    text = str(value or "").strip()
    if text in WIFI_CHOICES:
        return WIFI_CHOICES[text]
    compact = re.sub(r"\s+", "", text.lower())
    if compact in {"2.4&5", "2.4&5ghz", "wifi2.4&5", "wifi2.4&5ghz"}:
        return WIFI_CHOICES["2.4-5"]
    if compact in {
        "2.4&5&6",
        "2.4&5&6ghz",
        "2.4,5&6",
        "2.4,5&6ghz",
        "wifi2.4&5&6",
        "wifi2.4&5&6ghz",
    }:
        return WIFI_CHOICES["2.4-5-6"]
    if text in WIFI_CHOICES.values():
        return text
    valid = ", ".join([*WIFI_CHOICES.keys(), *WIFI_CHOICES.values()])
    raise ValueError(f"Unknown Wi-Fi choice: {value!r}. Valid choices: {valid}")


def sanitize_filename_part(value: str, fallback: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", str(value or "").strip()).strip("-._")
    return cleaned[:80] or fallback


def clone_cell_style(source, target) -> None:
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
    if source.hyperlink:
        target._hyperlink = copy(source.hyperlink)
    if source.comment:
        target.comment = copy(source.comment)


def replace_template_value(value, um: str, qsg: str, spec: str, wifi_name: str):
    if not isinstance(value, str):
        return value
    replaced = value.replace(OLD_UM, um).replace(OLD_QSG, qsg).replace(OLD_SPEC, spec)
    replaced = WIFI_PATTERN.sub(wifi_name, replaced)
    return ILLEGAL_XLSX_CHAR_PATTERN.sub("", replaced)


def update_xlsx_template(template_path: Path, output_path: Path, um: str, qsg: str, spec: str, wifi_name: str) -> None:
    wb = load_workbook(template_path)

    if "Cover" not in wb.sheetnames or "Interfaces" not in wb.sheetnames:
        raise ValueError(f"{template_path} does not contain required Cover and Interfaces sheets")

    cover = wb["Cover"]
    interfaces = wb["Interfaces"]
    cover["A88"] = um
    cover["B88"] = qsg
    cover["A90"] = spec
    interfaces["B3"] = wifi_name

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and not cell.value.startswith("="):
                    cell.value = replace_template_value(cell.value, um, qsg, spec, wifi_name)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def iter_xlsb_sheet_rows(workbook_path: Path, sheet_name: str):
    with open_xlsb(workbook_path) as wb:
        sheet = wb.get_sheet(sheet_name)
        for row in sheet.rows():
            yield row


def convert_xlsb_values_to_xlsx(template_path: Path, output_path: Path, um: str, qsg: str, spec: str, wifi_name: str) -> None:
    output_wb = Workbook()
    default_sheet = output_wb.active
    output_wb.remove(default_sheet)

    with open_xlsb(template_path) as input_wb:
        for sheet_name in input_wb.sheets:
            input_sheet = input_wb.get_sheet(sheet_name)
            ws = output_wb.create_sheet(title=sheet_name[:31])

            for row_index, row in enumerate(input_sheet.rows(), start=1):
                for column_index, cell in enumerate(row, start=1):
                    value = replace_template_value(cell.v, um, qsg, spec, wifi_name)
                    ws.cell(row=row_index, column=column_index, value=value)

            style_xlsb_output_sheet(ws)

    if "Cover" in output_wb.sheetnames:
        cover = output_wb["Cover"]
        cover["A88"] = um
        cover["B88"] = qsg
        cover["A90"] = spec
    if "Interfaces" in output_wb.sheetnames:
        output_wb["Interfaces"]["B3"] = wifi_name

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_wb.save(output_path)


def style_xlsb_output_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    title_fill = PatternFill("solid", fgColor="FCE4D6")
    header_font = Font(bold=True)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.row <= 2:
                cell.font = header_font
                cell.fill = title_fill if cell.row == 1 else header_fill

    for col_idx in range(1, min(ws.max_column, 18) + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for row_idx in range(1, min(ws.max_row, 40) + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, min(44, len(str(value).splitlines()[0])))
        ws.column_dimensions[letter].width = max_len + 2

    ws.freeze_panes = "A3"


def assert_inputs_exist(paths: Iterable[Path]) -> None:
    missing = [str(path) for path in paths if not path.exists()]
    if missing:
        raise FileNotFoundError("Missing template file(s): " + ", ".join(missing))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate RED EN 18031-1 and EN 18031-2 Excel workbooks.")
    parser.add_argument("--um", required=True, help="Cover tab UM value")
    parser.add_argument("--qsg", required=True, help="Cover tab QSG value")
    parser.add_argument("--spec", required=True, help="Cover tab Spec value or URL")
    parser.add_argument(
        "--wifi",
        required=True,
        help='Interface-01 Wi-Fi value. Use "2.4-5", "2.4-5-6", or the full display text.',
    )
    parser.add_argument("--en18031-1-template", type=Path, default=DEFAULT_EN18031_1_TEMPLATE)
    parser.add_argument("--en18031-2-template", type=Path, default=DEFAULT_EN18031_2_TEMPLATE)
    parser.add_argument("--output-dir", type=Path, default=Path("output/red-excel"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    wifi_name = normalize_wifi_choice(args.wifi)
    assert_inputs_exist([args.en18031_1_template, args.en18031_2_template])

    run_id = timestamp()
    model_part = sanitize_filename_part(args.um or args.qsg or "asus-router", "asus-router")
    output_dir = args.output_dir / f"{run_id}-{model_part}"
    en18031_1_output = output_dir / f"{run_id}-RED-DA-EN-18031-1-ASUS-Router.xlsx"
    en18031_2_output = output_dir / f"{run_id}-RED-18031-2-ASUS-Router.xlsx"

    update_xlsx_template(args.en18031_1_template, en18031_1_output, args.um, args.qsg, args.spec, wifi_name)
    convert_xlsb_values_to_xlsx(args.en18031_2_template, en18031_2_output, args.um, args.qsg, args.spec, wifi_name)

    print("Generated:")
    print(en18031_1_output)
    print(en18031_2_output)
    print()
    print("Notes:")
    print("- EN 18031-1 keeps the original .xlsx workbook formatting and formulas.")
    print("- EN 18031-2 is exported from .xlsb values into .xlsx; formulas/macros from the binary workbook are not preserved.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
